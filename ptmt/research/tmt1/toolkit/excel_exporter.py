from collections import defaultdict
from os import PathLike
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.cell import Cell
from openpyxl.styles import NamedStyle, PatternFill

from ptmt.research.dirs import DataDirectory


def color_correct(path: str | Path | PathLike):
    print("Correct coloring")
    print("loading excel!")
    wb = openpyxl.load_workbook(path)
    print("loaded excel")
    ws = wb[f"Ratings (ID_Only)"]

    col_iter = ws.columns
    next(col_iter)
    next(col_iter)
    c_iter = iter(next(col_iter))
    next(c_iter)
    next(c_iter)
    next(c_iter)


    row_iter = ws.rows
    next(row_iter)
    next(row_iter)
    next(row_iter)


    highlight = NamedStyle(name="highlight")
    highlight.fill = PatternFill("solid", bgColor="00CCFFCC", fgColor="00CCFFCC")
    wb.add_named_style(highlight)
    for c, row in zip(c_iter, row_iter):
        c: Cell
        row: tuple[Cell, ...]
        for value in row:
            if value.value == c.value:
                ws[value.coordinate].style = highlight
    wb.save(path)
    print("Correct coloring done!")



def export_excel(marker: str, data_dir: DataDirectory):
    fn = f'topic_models_{marker}.xlsx'
    if Path(fn).exists():
        print("Excel already exists!")
        return

    model = data_dir.load_original_py_model()
    translations = list(data_dir.iter_all_translations())
    with pd.ExcelWriter(fn, engine='openpyxl') as writer:
        for k in range(model.k):
            cols = []
            rows = defaultdict(list)
            cols.append(f'original')
            cols.append(f'P(original)')
            for i, value in enumerate(model.get_words_of_topic_sorted(k)[:20]):
                rows[f'Rank {i + 1}'].extend(value)
            for entry in translations:
                cols.append(f'{entry.path.name}')
                cols.append(f'P({entry.path.name})')
                for i, value in enumerate(entry.model_uncached.get_words_of_topic_sorted(k)[:20]):
                    rows[f'Rank {i + 1}'].extend(value)
            pd.DataFrame.from_dict(rows, orient='index', columns=cols).to_excel(writer, sheet_name=f"Topic {k}")

        # doc_id -> rank -> (topic_id, prob)

        m = defaultdict(list)
        col_names = dict()
        m_2 = defaultdict(list)
        col_names_2 = dict()

        for doc_id, v in data_dir.load_original_rating():
            for rank, value in enumerate(sorted(v, key=lambda x: x[1], reverse=True)):
                rank = rank + 1
                idx_entry = (doc_id, rank)
                col_names[('original', 'ID')] = None
                col_names[('original', 'P')] = None
                m[idx_entry].extend(value)
                col_names_2[('original', 'ID')] = None
                m_2[idx_entry].append(value[0])

        for entry in translations:
            for doc_id, v in entry.rating_uncached():
                for rank, value in enumerate(sorted(v, key=lambda x: x[1], reverse=True)):
                    rank = rank + 1
                    idx_entry = (doc_id, rank)
                    col_names[(entry.path.name, 'ID')] = None
                    col_names[(entry.path.name, 'P')] = None
                    m[idx_entry].extend(value)
                    col_names_2[(entry.path.name, 'ID')] = None
                    m_2[idx_entry].append(value[0])

        index = []
        columns = list(col_names.keys())
        data = []

        for k, dat in m.items():
            index.append(k)
            data.append(dat)
            assert len(dat) == len(columns), f'{len(dat)} != {len(columns)}'

        assert len(index) == len(data), f'{len(index)} != {len(data)}'

        index_names = ['doc_id', 'rank']
        column_names = ['origin', 'value']

        pd.DataFrame.from_dict(
            {
                'index': index,
                'columns': columns,
                'data': data,
                'index_names': index_names,
                'column_names': column_names
            },
            orient='tight',
        ).to_excel(writer, sheet_name=f"Ratings")

        index = []
        columns = list(col_names_2.keys())
        data = []

        for k, dat in m_2.items():
            index.append(k)
            data.append(dat)
            assert len(dat) == len(columns), f'{len(dat)} != {len(columns)}'

        assert len(index) == len(data), f'{len(index)} != {len(data)}'

        pd.DataFrame.from_dict(
            {
                'index': index,
                'columns': columns,
                'data': data,
                'index_names': index_names,
                'column_names': column_names
            },
            orient='tight',
        ).to_excel(writer, sheet_name=f"Ratings (ID_Only)")

    color_correct(fn)

